import openpyxl
import json

# Load the Excel file
wb = openpyxl.load_workbook('docs/legacy-forms/RLC Bingo Event Pull Tab Games Base Library.xlsx')
ws = wb.active

# Get column headers
headers = []
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(row=1, column=col).value
    headers.append(cell_value)

print("Column Headers:")
for i, header in enumerate(headers, 1):
    print(f"{i}: {header}")

print(f"\nTotal rows: {ws.max_row}")
print(f"Total columns: {ws.max_column}")

# Read first 5 data rows
print("\nFirst 5 data rows:")
for row in range(2, min(7, ws.max_row + 1)):  # Skip header row
    row_data = {}
    for col in range(1, ws.max_column + 1):
        header = headers[col-1]
        value = ws.cell(row=row, column=col).value
        row_data[header] = value
    print(f"Row {row}: {row_data}")

# Export all data to JSON for analysis
all_data = []
for row in range(2, ws.max_row + 1):  # Skip header row
    row_data = {}
    for col in range(1, ws.max_column + 1):
        header = headers[col-1]
        value = ws.cell(row=row, column=col).value
        row_data[header] = value
    all_data.append(row_data)

# Save to JSON file
with open('pulltabs_data.json', 'w') as f:
    json.dump(all_data, f, indent=2, default=str)

print(f"\nExported {len(all_data)} records to pulltabs_data.json")